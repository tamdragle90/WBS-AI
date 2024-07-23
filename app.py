# コード
import os
import pandas as pd
from datetime import datetime, timedelta
import holidays
import tkinter as tk
from tkinter import filedialog
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
def get_japanese_holidays(year):
    jp_holidays = holidays.Japan(years=year)
    holiday_list = [date for date in jp_holidays.keys()]
    return holiday_list
def read_and_preprocess_data(file_path, sheet_name):
    try:
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=7)
    except Exception as e:
        return None, None, None, None, None, None, None
    task_col_name = df.columns[1]
    assignee_col_name = df.columns[13]
    effort_col_name = df.columns[14]
    start_date_col_name = df.columns[28]
    end_date_col_name = df.columns[29]
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]
    start_date_value = ws['AE5'].value
    try:
        start_date = datetime.strptime(start_date_value, "%Y年%m月").date()
    except ValueError:
        return None, None, None, None, None, None, None
    return df, task_col_name, assignee_col_name, effort_col_name, start_date_col_name, end_date_col_name, start_date
def is_business_day(date, holidays_list):
    return date.weekday() < 5 and date not in holidays_list
def next_business_day(date, holidays_list):
    next_day = date + timedelta(days=1)
    while not is_business_day(next_day, holidays_list):
        next_day += timedelta(days=1)
    return next_day
def business_days_count(start_date, end_date, holidays):
    count = 0
    current_date = start_date
    while current_date <= end_date:
        if is_business_day(current_date, holidays):
            count += 1
        current_date += timedelta(days=1)
    return count
def validate_and_adjust_schedule(df, start_date, holidays_list, assignee_col_name, effort_col_name, start_date_col_name, end_date_col_name):
    error_rows = []
    user_workload = {}
    user_next_start_date = {}
    for index, row in df.iterrows():
        row_errors = []
        if pd.isna(row[effort_col_name]):
            continue 
        task_name_value = ""
        for col in df.columns[2:12]:
            if not pd.isna(row[col]):
                task_name_value = row[col]
                break
        if task_name_value == "":
            row_errors.append('Task name not null')
        if not pd.isna(row[start_date_col_name]) and not pd.isna(row[end_date_col_name]):  
            if row[start_date_col_name] > row[end_date_col_name]:
                row_errors.append('Start date is greater than end date')
            for date_col in [start_date_col_name, end_date_col_name]:
                try:
                    date = pd.to_datetime(row[date_col])
                    if not is_business_day(date, holidays_list):
                        row_errors.append(f'{date_col} is not a business day')
                except Exception as e:
                    row_errors.append(f'{date_col} is not a valid date: {e}')
        assignee = row[assignee_col_name]
        if pd.isna(assignee):
            assignee = min(user_workload, key=lambda k: sum(user_workload[k].values())) if user_workload else 'Unassigned'
            df.at[index, assignee_col_name] = assignee
        if assignee not in user_next_start_date:
            user_next_start_date[assignee] = pd.to_datetime(start_date)
            user_workload[assignee] = {}
        col_start_date = pd.to_datetime(row[start_date_col_name])
        col_end_date = pd.to_datetime(row[end_date_col_name])
        task_start_date = col_start_date
        if  pd.isna(row[start_date_col_name]) or task_start_date < user_next_start_date[assignee]:
            task_start_date = user_next_start_date[assignee]
        
        if task_start_date in user_workload[assignee]:
            while ((user_workload[assignee][task_start_date]) == 1):
                task_start_date = next_business_day(task_start_date, holidays_list)
                if(task_start_date not in user_workload[assignee]):
                    break
        if pd.isna(row[end_date_col_name]) or row[end_date_col_name] < task_start_date:
            col_end_date = task_start_date
        total_business_days = business_days_count(task_start_date, col_end_date, holidays_list)
        if total_business_days > 0:
            daily_effort = row[effort_col_name] / total_business_days
        else:
            daily_effort = row[effort_col_name]
        tmp_date = task_start_date
        tmp_end_date = col_end_date
        while tmp_date <= tmp_end_date:
            if is_business_day(tmp_date, holidays_list):
                if assignee not in user_next_start_date:
                    user_next_start_date[assignee] = start_date
                    user_workload[assignee] = {}
                remaining_effort = daily_effort
                date = tmp_date
                while remaining_effort > 0:
                    if date not in user_workload[assignee]:
                        user_workload[assignee][date] = 0
                    if user_workload[assignee][date] + remaining_effort <= 1:
                        user_workload[assignee][date] += remaining_effort
                        remaining_effort = 0
                    else:
                        available_effort = user_workload[assignee][date] + remaining_effort - 1
                        user_workload[assignee][date] = 1.0
                        remaining_effort = available_effort
                        date = next_business_day(date, holidays_list)
                        if (date > col_end_date):
                            row[end_date_col_name] = date
                            col_end_date = date
                        user_next_start_date[assignee] = date     
            tmp_date = next_business_day(tmp_date, holidays_list)     
        df.at[index, start_date_col_name] = task_start_.dt.strftime("%Y/%m/%d")
        df.at[index, end_date_col_name] = col_end_.dt.strftime("%Y/%m/%d")
        if row_errors:
            error_rows.append((index + 9, row_errors))
    return df, error_rows
from openpyxl.styles import NamedStyle, Font
def save_adjusted_schedule(df, input_file_path, output_file_path, sheet_name, error_rows):
    try:
        book = load_workbook(input_file_path)
        with pd.ExcelWriter(output_file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            sheet = writer.sheets[sheet_name]
            fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            for row, errors in error_rows:
                for cell in sheet[row]:
                    cell.fill = fill
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=8)
            writer.save()
    except Exception as e:
        print(f"エラー: {e}")
def main():
    input_file = "WBS.xlsx"
    sheet_name = 'スケジュール記入'
    current_year = datetime.now().year
    japanese_holidays = get_japanese_holidays(current_year)
    start_date = (datetime.now() + timedelta(days=1)).date()
    df, task_col_name, assignee_col_name, effort_col_name, start_date_col_name, end_date_col_name, start_date = read_and_preprocess_data(input_file, sheet_name)
    if df is None:
        return
    adjusted_df, error_rows = validate_and_adjust_schedule(df, start_date , japanese_holidays, assignee_col_name, effort_col_name, start_date_col_name, end_date_col_name)
    df.update(adjusted_df)
    directory = os.path.dirname(input_file)
    filename = os.path.basename(input_file)
    filename_without_extension, extension = os.path.splitext(filename)
    safe_filename = filename_without_extension.replace(" ", "_").replace("【", "").replace("】", "")
    output_filename = f"修正_{safe_filename}{extension}"
    output_file = os.path.join(directory, output_filename)
    save_adjusted_schedule(df, input_file, output_file, sheet_name, error_rows)
    print(f"{output_file}")
if __name__ == "__main__":
    main()